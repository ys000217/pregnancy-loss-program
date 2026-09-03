#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""EpiFactors panel catalog (SNV and SV counted separately) plus report figures.

Open this file in the IDE. Run with the local venv:

  D:\\ONT\\analyses\\burden\\.venv\\Scripts\\python.exe scripts\\epifactors_catalog.py

Defaults: rebuild tables from existing long TSVs if present, then plot.
Force a full ANNOVAR + VCF rescan with --catalog.

Rules (descriptive; no p-values):
  - genes = full EpiGenes_main.xlsx library
  - unit = gene (person x gene is 0/1)
  - SNV: drop synonymous; keep missense / LoF / splicing / other nonsyn
  - gnomAD AF is annotation; gene table also reports AF<1% subset
  - SV: drop mega events (>1 Mb or >20 genes); annotate ACMG and location
  - gene table splits SV into any-overlap / exon-or-splice / ACMG>=4
  - group rates are abnormal / normal / control proportions only
"""

from __future__ import annotations

import argparse
import csv
import gzip
import math
import os
import re
import shutil
import sys
from collections import defaultdict
from pathlib import Path

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
MODULE = Path(__file__).resolve().parents[1]
ROOT = Path(os.environ.get("BURDEN_ROOT", "D:/ONT/figure2"))
ONT = ROOT.parent
PHENO_PATH = ROOT / "sample_phenotype_648.tsv"
SNV_VCF = ROOT / "WGS_ONT_Intersection_648samples.vcf.gz"
SNV_ANNO = ROOT / "S956.snp.annovar.hg38_multianno.txt.gz"
SV_VCF = ONT / "clinical_649.GRCh38.correct.vcf"
SV_ANNO = ONT / "clinical_649.GRCh38.annotsv.tsv"
EPIGENES_XLSX = ROOT / "EpiGenes_main.xlsx"
OUT_DIR = Path(os.environ.get("BURDEN_OUT", str(MODULE / "tables"))) / "epifactors"
PLOT_DIR = Path(os.environ.get("BURDEN_PLOT", str(MODULE / "plots"))) / "epifactors"

GROUPS = ("abnormal", "normal", "control")
GROUP_COLOR = {"abnormal": "#D55E00", "normal": "#0072B2", "control": "#009E73"}
GROUP_LABEL = {"abnormal": "Abnormal", "normal": "Normal", "control": "Control"}
FUNC_KEEP = {"exonic", "splicing", "exonic;splicing"}
MAX_SV_BP = 1_000_000
MAX_SV_GENES = 20
POP_AF_RARE = 0.01
POP_AF_COLS = ["B_gain_AFmax", "B_loss_AFmax", "B_ins_AFmax", "B_inv_AFmax"]
MM = 1 / 25.4

GENES: set[str] = set()
GENE_CLASS: dict[str, str] = {}

REQUIRED_LONG = ("snv_variants.tsv", "sv_variants.tsv", "gene_panel.tsv")


# ---------------------------------------------------------------------------
# I/O helpers
# ---------------------------------------------------------------------------
def norm_chr(chrom: str) -> str:
    return chrom.lower().replace("chr", "").strip()


def open_text(path: Path):
    if str(path).endswith(".gz"):
        return gzip.open(path, "rt", encoding="utf-8", errors="replace")
    return open(path, "rt", encoding="utf-8", errors="replace")


def keep_gw_8_10(raw: str) -> bool:
    g = str(raw or "").strip().lower()
    return bool(
        re.fullmatch(r"g8|8(\+.*)?", g)
        or re.fullmatch(r"g9|9(\+.*)?", g)
        or re.fullmatch(r"g10|10(\+.*)?", g)
    )


def gw8_10_sample_ids(pheno_rows: list[dict]) -> set[str]:
    return {
        r["VCF_Sample_ID"]
        for r in pheno_rows
        if keep_gw_8_10(r.get("Gestational_Week", ""))
    }


def load_phenotype(path: Path) -> list[dict]:
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f, delimiter="\t"))


def has_alt_allele(gt_field: str) -> bool:
    gt = gt_field.split(":")[0]
    return any(ch in gt for ch in "123456789")


def parse_max_pop_af(row: dict) -> float | None:
    vals = []
    for col in POP_AF_COLS:
        try:
            v = float(row.get(col, "") or "nan")
        except ValueError:
            v = float("nan")
        if v == v:
            vals.append(v)
    return max(vals) if vals else None


def parse_acmg_class(raw: str) -> int | None:
    if raw is None:
        return None
    raw = str(raw).strip()
    if not raw or raw.lower() in {"na", "nan", "."}:
        return None
    m = re.match(r"^full=(\d+)$", raw)
    if m:
        return int(m.group(1))
    try:
        return int(float(raw))
    except ValueError:
        return None


def parse_info_field(info: str, key: str) -> str:
    for field in info.split(";"):
        if field.startswith(key + "="):
            return field.split("=", 1)[1]
    return ""


def parse_af(raw: str) -> float | None:
    if raw is None:
        return None
    raw = str(raw).strip()
    if raw in {"", ".", "NA", "na", "-"}:
        return None
    try:
        v = float(raw)
    except ValueError:
        return None
    return v if v == v else None


def write_tsv(path: Path, rows: list[dict], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, delimiter="\t", extrasaction="ignore")
        w.writeheader()
        for row in rows:
            w.writerow(row)


def read_tsv(path: Path) -> list[dict]:
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f, delimiter="\t"))


# ---------------------------------------------------------------------------
# Panel
# ---------------------------------------------------------------------------
def coarse_class(fun: str) -> str:
    fun = str(fun or "")
    if re.search(r"DNA modification", fun, re.I):
        return "dna_modification"
    if re.search(r"Chromatin remodeling", fun, re.I):
        return "chromatin_remodeling"
    if re.search(r"Polycomb group", fun, re.I):
        return "polycomb"
    if re.search(r"Histone chaperone", fun, re.I):
        return "histone_chaperone"
    if re.search(r"Histone modification erase", fun, re.I):
        return "histone_erase"
    if re.search(r"Histone modification write|Histone modification writer", fun, re.I):
        return "histone_write"
    if re.search(r"Histone modification read", fun, re.I):
        return "histone_read"
    if re.search(r"Histone modification", fun, re.I):
        return "histone_other"
    if re.search(r"RNA modification", fun, re.I):
        return "rna_modification"
    if re.search(r"(^|, )TF$|^TF$", fun):
        return "tf"
    return "other"


PRIO = {
    "dna_modification": 1,
    "chromatin_remodeling": 2,
    "polycomb": 3,
    "histone_chaperone": 4,
    "histone_erase": 5,
    "histone_write": 6,
    "histone_read": 7,
    "histone_other": 8,
    "rna_modification": 9,
    "tf": 10,
    "other": 11,
}


def load_panel_from_xlsx(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(x).strip() if x is not None else "" for x in next(rows_iter)]
    idx = {h: i for i, h in enumerate(header)}
    need = "HGNC_symbol"
    if need not in idx:
        raise SystemExit(f"{path} missing HGNC_symbol; columns={header[:12]}")
    seen: set[str] = set()
    out = []
    for raw in rows_iter:
        gene = str(raw[idx["HGNC_symbol"]] or "").strip().upper()
        if not gene or gene in {"#", "NAN", "NONE"} or gene in seen:
            continue
        seen.add(gene)
        fun = raw[idx["Function"]] if "Function" in idx else ""
        cls = coarse_class(fun)
        out.append(
            {
                "gene": gene,
                "class": cls,
                "priority": PRIO[cls],
                "epifactors_function": fun or "",
                "modification": raw[idx["Modification"]] if "Modification" in idx else "",
                "target": raw[idx["Target"]] if "Target" in idx else "",
                "complex_name": raw[idx["Complex_name"]] if "Complex_name" in idx else "",
                "source": "EpiGenes_main.xlsx",
                "match_rule": "full EpiFactors library (all HGNC symbols)",
            }
        )
    out.sort(key=lambda r: (r["priority"], r["gene"]))
    return out


def set_panel(panel_rows: list[dict]) -> None:
    global GENES, GENE_CLASS
    GENE_CLASS = {r["gene"].upper(): r.get("class") or r.get("gene_class") or "" for r in panel_rows}
    GENES = set(GENE_CLASS)


# ---------------------------------------------------------------------------
# Variant annotation
# ---------------------------------------------------------------------------
def gene_tokens(field: str) -> set[str]:
    if not field:
        return set()
    out = set()
    for tok in str(field).replace(",", ";").split(";"):
        tok = tok.strip()
        if not tok or tok in {".", "NA"}:
            continue
        tok = tok.split("(")[0].strip()
        if tok:
            out.add(tok.upper())
    return out


def hit_genes(field: str) -> list[str]:
    return sorted(g for g in gene_tokens(field) if g in GENES)


def snv_consequence(func: str, exonic: str) -> str:
    func = (func or "").strip()
    exonic = (exonic or "").strip().lower()
    if "splicing" in func:
        return "splicing"
    if any(x in exonic for x in ("stopgain", "stoploss", "frameshift", "startloss", "nonsense")):
        return "lof"
    if "missense" in exonic or "nonsynonymous" in exonic:
        return "missense"
    if "synonymous" in exonic or exonic in {"", ".", "unknown"}:
        return "skip"
    if exonic:
        return "other_nonsyn"
    return "skip"


def keep_sv_size(start: str, end: str, gene_field: str) -> bool:
    try:
        length = abs(int(float(end)) - int(float(start)))
    except (TypeError, ValueError):
        length = 0
    n_genes = max(1, len(gene_tokens(gene_field))) if gene_field else 1
    if length > MAX_SV_BP:
        return False
    if n_genes > MAX_SV_GENES:
        return False
    return True


def acmg_ge4(raw) -> int:
    try:
        return int(float(raw)) >= 4
    except (TypeError, ValueError):
        return 0


def sv_exon_or_splice(location: str) -> int:
    """AnnotSV uses 'exon1-exon1', not the word 'exonic'."""
    loc = (location or "").lower()
    if any(x in loc for x in ("splic", "cds")):
        return 1
    if "exon" in loc:
        return 1
    return 0


def location_class(location: str) -> str:
    loc = (location or "").lower()
    if sv_exon_or_splice(loc):
        return "exon_or_splice"
    if "intron" in loc:
        return "intron"
    if loc.strip() in {"", ".", "na"}:
        return "unannotated"
    return "other"


def collect_snv_anno(path: Path) -> list[dict]:
    rows = []
    with open_text(path) as f:
        header = f.readline().rstrip("\n").split("\t")
        idx = {c: i for i, c in enumerate(header)}
        for line in f:
            parts = line.rstrip("\n").split("\t")
            if len(parts) <= idx["ExonicFunc"]:
                continue
            gene_field = parts[idx["GeneName"]] if "GeneName" in idx else ""
            if not gene_field or gene_field in {".", "NA"}:
                gene_field = parts[idx["Gene"]]
            aa = parts[idx["AAChange"]] if "AAChange" in idx else ""
            genes = hit_genes(gene_field) or hit_genes(aa)
            if not genes:
                continue
            func = parts[idx["Func"]]
            if func not in FUNC_KEEP:
                continue
            if func == "exonic" and parts[idx["ExonicFunc"]] == "synonymous SNV":
                continue
            cons = snv_consequence(func, parts[idx["ExonicFunc"]])
            if cons == "skip":
                continue
            af = parse_af(parts[idx["GnomAD_genome_AF_ALL"]]) if "GnomAD_genome_AF_ALL" in idx else None
            rows.append(
                {
                    "chrom": parts[idx["CHROM"]],
                    "pos": parts[idx["POS"]],
                    "ref": parts[idx["REF"]],
                    "alt": parts[idx["ALT"]],
                    "key": (norm_chr(parts[idx["CHROM"]]), parts[idx["POS"]], parts[idx["REF"]], parts[idx["ALT"]]),
                    "genes": genes,
                    "func": func,
                    "exonic_func": parts[idx["ExonicFunc"]],
                    "aa_change": parts[idx["AAChange"]] if "AAChange" in idx else "",
                    "consequence": cons,
                    "gnomad_af": "" if af is None else af,
                    "pop_rare": int(af is None or af < POP_AF_RARE),
                    "clnsig": parts[idx["CLNSIG"]] if "CLNSIG" in idx else "",
                }
            )
    return rows


def attach_snv_carriers(anno_rows: list[dict], samples: list[str], vcf_path: Path) -> list[dict]:
    wanted = {r["key"]: r for r in anno_rows}
    for r in anno_rows:
        r["carriers"] = []
        r["gt_by_sample"] = {}
    with open_text(vcf_path) as f:
        vcf_samples = None
        for line in f:
            if line.startswith("#CHROM"):
                vcf_samples = line.strip().split("\t")[9:]
                continue
            if line.startswith("#"):
                continue
            parts = line.rstrip("\n").split("\t")
            if parts[6] != "PASS":
                continue
            rec = wanted.get((norm_chr(parts[0]), parts[1], parts[3], parts[4]))
            if rec is None:
                continue
            gt_map = dict(zip(vcf_samples, parts[9:]))
            carriers = []
            for s in samples:
                gt = gt_map.get(s, "./.")
                if has_alt_allele(gt):
                    carriers.append(s)
                    rec["gt_by_sample"][s] = gt.split(":")[0]
            rec["carriers"] = carriers
    return anno_rows


def collect_sv_gene_hits(anno_path: Path) -> dict[str, dict]:
    hits: dict[str, dict] = {}
    with open(anno_path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            genes = hit_genes(row.get("Gene_name", ""))
            if not genes:
                continue
            chrom = row.get("SV_chrom")
            start = row.get("SV_start")
            end = row.get("SV_end")
            svtype = row.get("SV_type")
            if not keep_sv_size(start, end, row.get("Gene_name", "")):
                continue
            coord_key = f"{norm_chr(chrom)}:{start}:{end}:{svtype}"
            mode = row.get("Annotation_mode", "")
            loc = row.get("Location", "")
            acmg = parse_acmg_class(row.get("ACMG_class"))
            score = float(row.get("AnnotSV_ranking_score") or 0)
            rec = {
                "coord_key": coord_key,
                "chrom": chrom,
                "start": start,
                "end": end,
                "svtype": svtype,
                "sv_id": row.get("ID"),
                "genes": genes,
                "location": loc,
                "mode": mode,
                "acmg_class": acmg if acmg is not None else "",
                "ranking_score": score,
                "max_pop_af": parse_max_pop_af(row),
            }
            prev = hits.get(coord_key)
            if prev is None:
                hits[coord_key] = rec
                continue
            prev_split = prev.get("mode") == "split"
            new_split = mode == "split"
            if new_split and not prev_split:
                hits[coord_key] = rec
            elif new_split == prev_split and score > prev.get("ranking_score", 0):
                hits[coord_key] = rec
            else:
                extra = set(prev["genes"]) | set(genes)
                prev["genes"] = sorted(extra)
                if loc and loc not in str(prev.get("location", "")):
                    prev["location"] = f"{prev.get('location')};{loc}"
    return hits


def attach_sv_carriers(sv_hits: dict[str, dict], samples: list[str], vcf_path: Path) -> list[dict]:
    with open_text(vcf_path) as f:
        vcf_samples = None
        for line in f:
            if line.startswith("#CHROM"):
                vcf_samples = line.strip().split("\t")[9:]
                continue
            if line.startswith("#"):
                continue
            parts = line.rstrip("\n").split("\t")
            if parts[6] != "PASS":
                continue
            svtype = parse_info_field(parts[7], "SVTYPE") or "UNK"
            sv_end = parse_info_field(parts[7], "END") or parts[1]
            rec = sv_hits.get(f"{norm_chr(parts[0])}:{parts[1]}:{sv_end}:{svtype}")
            if rec is None:
                continue
            gt_map = dict(zip(vcf_samples, parts[9:]))
            carriers = []
            gts = {}
            for s in samples:
                gt = gt_map.get(s, "./.")
                if has_alt_allele(gt):
                    carriers.append(s)
                    gts[s] = gt.split(":")[0]
            rec["carriers"] = carriers
            rec["gt_by_sample"] = gts
            rec["sv_id"] = rec.get("sv_id") or parts[2]
    return [r for r in sv_hits.values() if r.get("carriers")]


def group_counts(carriers: list[str], groups: dict[str, list[str]]) -> dict:
    cset = set(carriers)
    out = {}
    for name, members in groups.items():
        n = sum(1 for s in members if s in cset)
        out[f"{name}_n"] = n
        out[f"{name}_rate"] = n / len(members) if members else 0.0
    return out


# ---------------------------------------------------------------------------
# Aggregate tables
# ---------------------------------------------------------------------------
def count_by_gene(rows: list[dict], groups: dict[str, list[str]], pred) -> dict[str, dict]:
    hits = defaultdict(lambda: defaultdict(set))
    meta = {}
    for r in rows:
        if not pred(r):
            continue
        gene = r["gene"]
        hits[gene][r.get("Condition", "")].add(r["Sample_ID"])
        meta[gene] = r.get("gene_class", "")
    n = {g: len(members) for g, members in groups.items()}
    out = {}
    for gene, by_cond in hits.items():
        row = {"gene": gene, "gene_class": meta.get(gene, GENE_CLASS.get(gene, ""))}
        for g in GROUPS:
            k = len(by_cond.get(g, set()))
            row[f"{g}_n"] = k
            row[f"{g}_rate"] = k / n[g] if n.get(g) else 0.0
        row["rate_ab_minus_normal"] = row["abnormal_rate"] - row["normal_rate"]
        row["rate_ab_minus_control"] = row["abnormal_rate"] - row["control_rate"]
        out[gene] = row
    return out


def build_snv_gene_rows(snv_long: list[dict], groups: dict[str, list[str]]) -> list[dict]:
    snv_nonsyn = count_by_gene(snv_long, groups, lambda r: True)
    snv_rare = count_by_gene(snv_long, groups, lambda r: str(r.get("pop_rare")) in {"1", "1.0", 1})
    rows = []
    for gene in set(snv_nonsyn) | set(snv_rare):
        a = snv_nonsyn.get(gene, {})
        b = snv_rare.get(gene, {})
        row = {"gene": gene, "gene_class": GENE_CLASS.get(gene, a.get("gene_class") or b.get("gene_class") or "")}
        for prefix, rec in (("nonsyn", a), ("rare", b)):
            for g in GROUPS:
                row[f"{prefix}_{g}_n"] = rec.get(f"{g}_n", 0)
                row[f"{prefix}_{g}_rate"] = rec.get(f"{g}_rate", 0.0)
            row[f"{prefix}_rate_ab_minus_normal"] = rec.get("rate_ab_minus_normal", 0.0)
            row[f"{prefix}_rate_ab_minus_control"] = rec.get("rate_ab_minus_control", 0.0)
        rows.append(row)
    rows.sort(key=lambda r: (-float(r["nonsyn_abnormal_rate"]), -int(r["nonsyn_abnormal_n"]), r["gene"]))
    return rows


def build_sv_gene_rows(sv_long: list[dict], groups: dict[str, list[str]]) -> list[dict]:
    sv_any = count_by_gene(sv_long, groups, lambda r: True)
    sv_exon = count_by_gene(sv_long, groups, lambda r: int(r.get("exon_or_splice") or 0) == 1)
    sv_p = count_by_gene(sv_long, groups, lambda r: int(r.get("acmg_ge4") or 0) == 1)
    rows = []
    for gene in set(sv_any) | set(sv_exon) | set(sv_p):
        row = {"gene": gene, "gene_class": GENE_CLASS.get(gene, "")}
        for prefix, rec in (
            ("any", sv_any.get(gene, {})),
            ("exon", sv_exon.get(gene, {})),
            ("acmg4", sv_p.get(gene, {})),
        ):
            for g in GROUPS:
                row[f"{prefix}_{g}_n"] = rec.get(f"{g}_n", 0)
                row[f"{prefix}_{g}_rate"] = rec.get(f"{g}_rate", 0.0)
            row[f"{prefix}_rate_ab_minus_normal"] = rec.get("rate_ab_minus_normal", 0.0)
            row[f"{prefix}_rate_ab_minus_control"] = rec.get("rate_ab_minus_control", 0.0)
        rows.append(row)
    rows.sort(key=lambda r: (-float(r["any_abnormal_rate"]), -float(r["exon_abnormal_rate"]), r["gene"]))
    return rows


SNV_GENE_FIELDS = [
    "gene", "gene_class",
    "nonsyn_abnormal_n", "nonsyn_abnormal_rate", "nonsyn_normal_n", "nonsyn_normal_rate",
    "nonsyn_control_n", "nonsyn_control_rate",
    "nonsyn_rate_ab_minus_normal", "nonsyn_rate_ab_minus_control",
    "rare_abnormal_n", "rare_abnormal_rate", "rare_normal_n", "rare_normal_rate",
    "rare_control_n", "rare_control_rate",
    "rare_rate_ab_minus_normal", "rare_rate_ab_minus_control",
]
SV_GENE_FIELDS = [
    "gene", "gene_class",
    "any_abnormal_n", "any_abnormal_rate", "any_normal_n", "any_normal_rate",
    "any_control_n", "any_control_rate",
    "any_rate_ab_minus_normal", "any_rate_ab_minus_control",
    "exon_abnormal_n", "exon_abnormal_rate", "exon_normal_n", "exon_normal_rate",
    "exon_control_n", "exon_control_rate",
    "exon_rate_ab_minus_normal", "exon_rate_ab_minus_control",
    "acmg4_abnormal_n", "acmg4_abnormal_rate", "acmg4_normal_n", "acmg4_normal_rate",
    "acmg4_control_n", "acmg4_control_rate",
    "acmg4_rate_ab_minus_normal", "acmg4_rate_ab_minus_control",
]
CARD_FIELDS = [
    "Sample_ID", "Condition", "n_snv_nonsyn", "n_snv_rare", "n_sv_any", "n_sv_exon", "n_sv_acmg4",
    "snv_genes_nonsyn", "snv_genes_rare", "sv_genes_any", "sv_genes_exon", "sv_genes_acmg4",
    "snv_variant_pos", "snv_rare_variant_pos", "sv_variant_pos",
]


def build_cards(snv_long: list[dict], sv_long: list[dict], abnormal: list[str]) -> list[dict]:
    cards = []
    for s in abnormal:
        snv_all = [r for r in snv_long if r["Sample_ID"] == s]
        snv_rare_s = [r for r in snv_all if str(r.get("pop_rare")) in {"1", "1.0", 1}]
        sv_all = [r for r in sv_long if r["Sample_ID"] == s]
        sv_ex = [r for r in sv_all if int(float(r.get("exon_or_splice") or 0)) == 1]
        sv_ac = [r for r in sv_all if int(float(r.get("acmg_ge4") or 0)) == 1]
        cards.append(
            {
                "Sample_ID": s,
                "Condition": "abnormal",
                "n_snv_nonsyn": len(snv_all),
                "n_snv_rare": len(snv_rare_s),
                "n_sv_any": len(sv_all),
                "n_sv_exon": len(sv_ex),
                "n_sv_acmg4": len(sv_ac),
                "snv_genes_nonsyn": ";".join(sorted({r["gene"] for r in snv_all})),
                "snv_genes_rare": ";".join(sorted({r["gene"] for r in snv_rare_s})),
                "sv_genes_any": ";".join(sorted({r["gene"] for r in sv_all})),
                "sv_genes_exon": ";".join(sorted({r["gene"] for r in sv_ex})),
                "sv_genes_acmg4": ";".join(sorted({r["gene"] for r in sv_ac})),
                "snv_variant_pos": " | ".join(
                    f"{r['gene']}:{r['consequence']}:{r.get('variant_pos', '')}" for r in snv_all
                ),
                "snv_rare_variant_pos": " | ".join(
                    f"{r['gene']}:{r['consequence']}:{r.get('variant_pos', '')}" for r in snv_rare_s
                ),
                "sv_variant_pos": " | ".join(
                    f"{r['gene']}:{r.get('svtype')}:{r.get('variant_pos')}:ACMG{r.get('acmg_class')}:{r.get('location')}"
                    for r in sv_all
                ),
            }
        )
    return cards


def per_sample_metrics(snv_long: list[dict], sv_long: list[dict], samples: list[str], pheno_map: dict) -> list[dict]:
    snv = defaultdict(list)
    rare = defaultdict(list)
    sv = defaultdict(list)
    sv_ex = defaultdict(list)
    for r in snv_long:
        snv[r["Sample_ID"]].append(r)
        if str(r.get("pop_rare")) in {"1", "1.0", 1}:
            rare[r["Sample_ID"]].append(r)
    for r in sv_long:
        sv[r["Sample_ID"]].append(r)
        if int(float(r.get("exon_or_splice") or 0)) == 1:
            sv_ex[r["Sample_ID"]].append(r)
    rows = []
    for s in samples:
        rows.append(
            {
                "Sample_ID": s,
                "Condition": pheno_map.get(s, ""),
                "n_snv_nonsyn": len(snv[s]),
                "n_snv_rare": len(rare[s]),
                "n_sv_any": len(sv[s]),
                "n_sv_exon": len(sv_ex[s]),
            }
        )
    return rows


# ---------------------------------------------------------------------------
# Catalog
# ---------------------------------------------------------------------------
def run_catalog(out: Path, panel_rows: list[dict]) -> tuple[list[dict], list[dict]]:
    set_panel(panel_rows)
    write_tsv(out / "gene_panel.tsv", panel_rows, list(panel_rows[0].keys()))

    pheno_rows = load_phenotype(PHENO_PATH)
    pheno_map = {r["VCF_Sample_ID"]: r["Condition"] for r in pheno_rows}
    samples = [r["VCF_Sample_ID"] for r in pheno_rows]
    groups = {g: [s for s in samples if pheno_map.get(s) == g] for g in GROUPS}
    print(
        f"panel={len(GENES)} abnormal={len(groups['abnormal'])} "
        f"normal={len(groups['normal'])} control={len(groups['control'])}"
    )

    print("[1/3] ANNOVAR (drop synonymous)...")
    snv_anno = collect_snv_anno(SNV_ANNO)
    print(f"  non-synonymous/splice sites in panel: {len(snv_anno)}")
    print("[2/3] SNV genotypes...")
    snv_anno = attach_snv_carriers(snv_anno, samples, SNV_VCF)
    snv_anno = [r for r in snv_anno if r.get("carriers")]
    print(f"  sites with >=1 carrier: {len(snv_anno)}")

    snv_long, snv_site = [], []
    for rec in snv_anno:
        gc = group_counts(rec["carriers"], groups)
        for gene in rec["genes"]:
            site = {
                "chrom": rec["chrom"], "pos": rec["pos"], "ref": rec["ref"], "alt": rec["alt"],
                "gene": gene, "gene_class": GENE_CLASS.get(gene, ""),
                "consequence": rec["consequence"], "func": rec["func"],
                "exonic_func": rec["exonic_func"], "aa_change": rec["aa_change"],
                "gnomad_af": rec["gnomad_af"], "pop_rare": rec["pop_rare"],
                "clnsig": rec["clnsig"], "n_carriers": len(rec["carriers"]), **gc,
            }
            snv_site.append(site)
            for s in rec["carriers"]:
                snv_long.append(
                    {
                        "Sample_ID": s, "Condition": pheno_map.get(s, ""),
                        "chrom": rec["chrom"], "pos": rec["pos"], "ref": rec["ref"], "alt": rec["alt"],
                        "gt": rec.get("gt_by_sample", {}).get(s, ""),
                        "gene": gene, "gene_class": GENE_CLASS.get(gene, ""),
                        "consequence": rec["consequence"], "exonic_func": rec["exonic_func"],
                        "aa_change": rec["aa_change"], "gnomad_af": rec["gnomad_af"],
                        "pop_rare": rec["pop_rare"],
                        "variant_pos": f"{rec['chrom']}:{rec['pos']}:{rec['ref']}>{rec['alt']}",
                    }
                )
    write_tsv(
        out / "snv_variants.tsv",
        snv_long,
        [
            "Sample_ID", "Condition", "gene", "gene_class", "chrom", "pos", "ref", "alt",
            "variant_pos", "gt", "consequence", "exonic_func", "aa_change", "gnomad_af", "pop_rare",
        ],
    )
    write_tsv(
        out / "snv_sites.tsv",
        snv_site,
        [
            "chrom", "pos", "ref", "alt", "gene", "gene_class", "consequence", "func",
            "exonic_func", "aa_change", "gnomad_af", "pop_rare", "clnsig", "n_carriers",
            "abnormal_n", "abnormal_rate", "normal_n", "normal_rate", "control_n", "control_rate",
        ],
    )

    print("[3/3] AnnotSV + SV VCF...")
    sv_hits = collect_sv_gene_hits(SV_ANNO)
    print(f"  size-filtered loci overlapping panel: {len(sv_hits)}")
    sv_hits = attach_sv_carriers(sv_hits, samples, SV_VCF)
    print(f"  loci with >=1 carrier: {len(sv_hits)}")

    sv_long, sv_site = [], []
    for rec in sv_hits:
        try:
            svlen = abs(int(float(rec["end"])) - int(float(rec["start"])))
        except (TypeError, ValueError):
            svlen = ""
        gc = group_counts(rec["carriers"], groups)
        loc = rec.get("location") or ""
        exon = sv_exon_or_splice(loc)
        acmg = rec.get("acmg_class")
        acmg4 = acmg_ge4(acmg)
        chrom, start, end = rec.get("chrom"), rec.get("start"), rec.get("end")
        vpos = f"chr{norm_chr(str(chrom))}:{start}-{end}:{rec.get('svtype')}"
        for gene in rec["genes"]:
            sv_site.append(
                {
                    "coord_key": rec["coord_key"], "chrom": chrom, "start": start, "end": end,
                    "svlen": svlen, "svtype": rec.get("svtype"), "sv_id": rec.get("sv_id"),
                    "variant_pos": vpos, "gene": gene, "gene_class": GENE_CLASS.get(gene, ""),
                    "location": loc, "exon_or_splice": exon, "acmg_class": acmg, "acmg_ge4": acmg4,
                    "max_pop_af": rec.get("max_pop_af") if rec.get("max_pop_af") is not None else "",
                    "n_carriers": len(rec["carriers"]), **gc,
                }
            )
            for s in rec["carriers"]:
                sv_long.append(
                    {
                        "Sample_ID": s, "Condition": pheno_map.get(s, ""),
                        "coord_key": rec["coord_key"], "sv_id": rec.get("sv_id"),
                        "chrom": chrom, "start": start, "end": end, "svlen": svlen,
                        "svtype": rec.get("svtype"), "variant_pos": vpos,
                        "gt": rec.get("gt_by_sample", {}).get(s, ""),
                        "gene": gene, "gene_class": GENE_CLASS.get(gene, ""),
                        "location": loc, "exon_or_splice": exon, "acmg_class": acmg, "acmg_ge4": acmg4,
                    }
                )
    write_tsv(
        out / "sv_variants.tsv",
        sv_long,
        [
            "Sample_ID", "Condition", "gene", "gene_class", "chrom", "start", "end",
            "variant_pos", "svtype", "svlen", "sv_id", "gt", "location",
            "exon_or_splice", "acmg_class", "acmg_ge4",
        ],
    )
    write_tsv(
        out / "sv_sites.tsv",
        sv_site,
        [
            "coord_key", "chrom", "start", "end", "svlen", "svtype", "sv_id", "variant_pos",
            "gene", "gene_class", "location", "exon_or_splice", "acmg_class", "acmg_ge4",
            "max_pop_af", "n_carriers", "abnormal_n", "abnormal_rate",
            "normal_n", "normal_rate", "control_n", "control_rate",
        ],
    )
    return snv_long, sv_long


def refresh_derived(
    out: Path,
    snv_long: list[dict],
    sv_long: list[dict],
    keep_ids: set[str] | None = None,
    pheno_path: Path | None = None,
) -> None:
    for r in sv_long:
        r["exon_or_splice"] = sv_exon_or_splice(r.get("location", ""))
        r["acmg_ge4"] = acmg_ge4(r.get("acmg_class"))
        r["loc_class"] = location_class(r.get("location", ""))
    pheno_rows = load_phenotype(pheno_path or PHENO_PATH)
    if keep_ids is not None:
        pheno_rows = [r for r in pheno_rows if r["VCF_Sample_ID"] in keep_ids]
        snv_long = [r for r in snv_long if r["Sample_ID"] in keep_ids]
        sv_long = [r for r in sv_long if r["Sample_ID"] in keep_ids]
    pheno_map = {r["VCF_Sample_ID"]: r["Condition"] for r in pheno_rows}
    samples = [r["VCF_Sample_ID"] for r in pheno_rows]
    groups = {g: [s for s in samples if pheno_map.get(s) == g] for g in GROUPS}

    snv_gene = build_snv_gene_rows(snv_long, groups)
    sv_gene = build_sv_gene_rows(sv_long, groups)
    cards = build_cards(snv_long, sv_long, groups["abnormal"])
    metrics = per_sample_metrics(snv_long, sv_long, samples, pheno_map)
    write_tsv(out / "snv_by_gene.tsv", snv_gene, SNV_GENE_FIELDS)
    write_tsv(out / "sv_by_gene.tsv", sv_gene, SV_GENE_FIELDS)
    write_tsv(out / "abnormal_sample_cards.tsv", cards, CARD_FIELDS)
    write_tsv(
        out / "sample_metrics.tsv",
        metrics,
        ["Sample_ID", "Condition", "n_snv_nonsyn", "n_snv_rare", "n_sv_any", "n_sv_exon"],
    )
    print(f"SNV gene rows={len(snv_gene)} SV gene rows={len(sv_gene)} cards={len(cards)}")


# ---------------------------------------------------------------------------
# Figures (Nature Portfolio: 180 mm, Arial 5–7 pt, Okabe–Ito, vector PDF)
# ---------------------------------------------------------------------------
def apply_nature_style() -> None:
    import matplotlib as mpl

    mpl.rcParams.update(
        {
            "font.family": "Arial",
            "font.size": 7,
            "axes.labelsize": 7,
            "xtick.labelsize": 6,
            "ytick.labelsize": 6,
            "legend.fontsize": 6,
            "axes.linewidth": 0.6,
            "xtick.major.width": 0.6,
            "ytick.major.width": 0.6,
            "xtick.major.size": 2.5,
            "ytick.major.size": 2.5,
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
            "axes.unicode_minus": False,
            "figure.dpi": 300,
            "savefig.dpi": 600,
            "axes.spines.top": False,
            "axes.spines.right": False,
        }
    )


def save_fig(fig, stem: str) -> None:
    PLOT_DIR.mkdir(parents=True, exist_ok=True)
    pdf = PLOT_DIR / f"{stem}.pdf"
    png = PLOT_DIR / f"{stem}.png"
    fig.savefig(png, facecolor="white")
    try:
        fig.savefig(pdf, facecolor="white")
    except PermissionError:
        alt = PLOT_DIR / f"{stem}_new.pdf"
        fig.savefig(alt, facecolor="white")
        print(f"  PDF locked, wrote {alt.name}")
    print(f"  wrote {pdf.name} / {png.name}")


def panel_label(ax, letter: str) -> None:
    ax.text(
        -0.12, 1.08, letter, transform=ax.transAxes, fontsize=8, fontweight="bold",
        va="bottom", ha="left", color="black", clip_on=False,
    )


def mannwhitney_p(x, y) -> float:
    """Two-sided Mann–Whitney U (normal approximation, tie correction)."""
    x = [float(v) for v in x]
    y = [float(v) for v in y]
    n1, n2 = len(x), len(y)
    if n1 < 1 or n2 < 1:
        return 1.0
    combined = [(v, 0) for v in x] + [(v, 1) for v in y]
    combined.sort(key=lambda t: t[0])
    ranks = [0.0] * len(combined)
    tie_term = 0.0
    i, n = 0, len(combined)
    while i < n:
        j = i
        while j < n and combined[j][0] == combined[i][0]:
            j += 1
        avg = 0.5 * (i + 1 + j)
        t = j - i
        if t > 1:
            tie_term += t * t * t - t
        for k in range(i, j):
            ranks[k] = avg
        i = j
    r1 = sum(ranks[k] for k, (_, g) in enumerate(combined) if g == 0)
    u1 = r1 - n1 * (n1 + 1) / 2.0
    mu = n1 * n2 / 2.0
    ntot = n1 + n2
    if ntot <= 1:
        return 1.0
    sigma2 = n1 * n2 / 12.0 * (ntot + 1 - (tie_term / (ntot * (ntot - 1)) if ntot > 1 else 0))
    if sigma2 <= 0:
        return 1.0
    z = abs(u1 - mu) - 0.5
    z = max(z, 0.0) / (sigma2 ** 0.5)
    return min(1.0, math.erfc(z / (2 ** 0.5)))


def fmt_p(p: float) -> str:
    if p < 0.001:
        return f"{p:.1e}"
    return f"{p:.3f}"


def annotate_mw(ax, data: dict[str, list[float]]) -> dict[str, float]:
    """Write pairwise two-sided Mann–Whitney p-values above the axes."""
    pairs = (
        ("abnormal", "normal", "Ab–N"),
        ("abnormal", "control", "Ab–C"),
        ("normal", "control", "N–C"),
    )
    ps = {}
    bits = []
    for a, b, lab in pairs:
        p = mannwhitney_p(data[a], data[b])
        ps[f"{a}_vs_{b}"] = p
        bits.append(f"{lab} p={fmt_p(p)}")
    ax.text(
        0.5, 1.01, "Mann–Whitney " + "; ".join(bits),
        transform=ax.transAxes, ha="center", va="bottom", fontsize=5, color="black",
    )
    return ps


def box_with_points(
    ax,
    data: dict[str, list[float]],
    ylabel: str,
    group_ns: dict[str, int] | None = None,
) -> None:
    import numpy as np

    rng = np.random.default_rng(1)
    positions = list(range(1, len(GROUPS) + 1))
    series = [data[g] for g in GROUPS]
    bp = ax.boxplot(
        series, positions=positions, widths=0.55, showfliers=False, patch_artist=True,
        medianprops={"color": "black", "linewidth": 0.8},
        whiskerprops={"linewidth": 0.6, "color": "black"},
        capprops={"linewidth": 0.6, "color": "black"},
        boxprops={"linewidth": 0.6, "color": "black"},
    )
    for patch, g in zip(bp["boxes"], GROUPS):
        patch.set_facecolor(GROUP_COLOR[g])
        patch.set_alpha(0.35)
    for i, g in enumerate(GROUPS, start=1):
        y = np.asarray(data[g], dtype=float)
        x = np.full(len(y), i, dtype=float) + rng.uniform(-0.12, 0.12, len(y))
        ax.scatter(x, y, s=5, c=GROUP_COLOR[g], alpha=0.45, linewidths=0, zorder=3, rasterized=False)
    ax.set_xticks(positions)
    if group_ns:
        ax.set_xticklabels([f"{GROUP_LABEL[g]}\n(n={group_ns[g]})" for g in GROUPS])
    else:
        ax.set_xticklabels([GROUP_LABEL[g] for g in GROUPS])
    ax.set_ylabel(ylabel)


def fisher_ab_ctrl(ab_k: int, ab_n: int, ctrl_k: int, ctrl_n: int) -> float:
    from compute_burden import fisher_exact_2x2

    return fisher_exact_2x2(ab_k, ab_n - ab_k, ctrl_k, ctrl_n - ctrl_k)


def sv_unique_site_rates(sv_long: list[dict], groups: dict[str, list[str]]) -> list[dict]:
    """One row per unique SV (chrom-start-end-type), gene-overlap collapsed."""
    from compute_burden import bh_fdr

    hits: dict[str, dict[str, set[str]]] = defaultdict(lambda: {g: set() for g in GROUPS})
    meta: dict[str, dict] = {}
    for r in sv_long:
        key = f"{norm_chr(r.get('chrom', ''))}:{r.get('start')}:{r.get('end')}:{r.get('svtype', '')}"
        cond = r.get("Condition", "")
        if cond in hits[key]:
            hits[key][cond].add(r["Sample_ID"])
        if key not in meta:
            meta[key] = r
    n = {g: len(members) for g, members in groups.items()}
    rows = []
    for key, by in hits.items():
        r0 = meta[key]
        row = {
            "coord_key": key,
            "svtype": r0.get("svtype", ""),
            "acmg_class": r0.get("acmg_class", ""),
        }
        for g in GROUPS:
            k = len(by[g])
            row[f"{g}_n"] = k
            row[f"{g}_rate"] = k / n[g] if n[g] else 0.0
        row["p_abnormal_vs_control"] = fisher_ab_ctrl(
            int(row["abnormal_n"]), n["abnormal"], int(row["control_n"]), n["control"]
        )
        row["abnormal_rate"] = row["abnormal_rate"]
        row["control_rate"] = row["control_rate"]
        row["normal_rate"] = row["normal_rate"]
        rows.append(row)
    if rows:
        qvals = bh_fdr([r["p_abnormal_vs_control"] for r in rows])
        for r, q in zip(rows, qvals):
            r["fdr_abnormal_vs_control"] = q
    rows.sort(key=lambda r: r["p_abnormal_vs_control"])
    return rows


def scatter_ab_vs_ctrl(ax, rows: list[dict], fdr_col: str = "fdr_abnormal_vs_control") -> None:
    xs = [100 * fnum(r["control_rate"]) for r in rows]
    ys = [100 * fnum(r["abnormal_rate"]) for r in rows]
    sig = [fnum(r.get(fdr_col, 1.0)) < 0.05 for r in rows]
    ax.plot([0, 100], [0, 100], color="#888888", lw=0.5, ls="--")
    ns_x = [x for x, s in zip(xs, sig) if not s]
    ns_y = [y for y, s in zip(ys, sig) if not s]
    s_x = [x for x, s in zip(xs, sig) if s]
    s_y = [y for y, s in zip(ys, sig) if s]
    ax.scatter(ns_x, ns_y, c="#0072B2", s=9, linewidths=0, alpha=0.75, zorder=2)
    if s_x:
        ax.scatter(s_x, s_y, c="#D55E00", s=12, linewidths=0, alpha=0.9, zorder=3)
    ax.set_xlabel("Control carrier rate (%)")
    ax.set_ylabel("Abnormal carrier rate (%)")
    ax.set_xlim(-2, 102)
    ax.set_ylim(-2, 102)
    n_sig = sum(sig)
    ax.text(
        0.04, 0.96, f"n={len(rows)} sites; FDR<0.05: {n_sig}",
        transform=ax.transAxes, va="top", ha="left", fontsize=5.5,
    )


def write_source(name: str, rows: list[dict], fields: list[str]) -> None:
    src = PLOT_DIR / "source_data"
    write_tsv(src / name, rows, fields)


def fnum(x, default=0.0) -> float:
    try:
        return float(x)
    except (TypeError, ValueError):
        return default


def inum(x, default=0) -> int:
    try:
        return int(float(x))
    except (TypeError, ValueError):
        return default


GENE_RANK_FIELDS = [
    "gene", "gene_class",
    "abnormal_n", "abnormal_rate", "normal_n", "normal_rate", "control_n", "control_rate",
    "delta_rate_ab_ctrl", "p_abnormal_vs_control",
]


def rank_genes_ab_vs_control(
    gene_rows: list[dict],
    prefix: str,
    n: dict[str, int],
) -> list[dict]:
    """Rank genes by raw Fisher p (abnormal vs control carriers); smallest p first."""
    ranked = []
    for r in gene_rows:
        ab_n = inum(r.get(f"{prefix}_abnormal_n"))
        ctrl_n = inum(r.get(f"{prefix}_control_n"))
        ab_rate = fnum(r.get(f"{prefix}_abnormal_rate"))
        ctrl_rate = fnum(r.get(f"{prefix}_control_rate"))
        ranked.append(
            {
                "gene": r["gene"],
                "gene_class": r.get("gene_class", GENE_CLASS.get(r["gene"], "")),
                "abnormal_n": ab_n,
                "abnormal_rate": ab_rate,
                "normal_n": inum(r.get(f"{prefix}_normal_n")),
                "normal_rate": fnum(r.get(f"{prefix}_normal_rate")),
                "control_n": ctrl_n,
                "control_rate": ctrl_rate,
                "delta_rate_ab_ctrl": ab_rate - ctrl_rate,
                "p_abnormal_vs_control": fisher_ab_ctrl(ab_n, n["abnormal"], ctrl_n, n["control"]),
            }
        )
    ranked.sort(
        key=lambda x: (
            x["p_abnormal_vs_control"],
            -abs(x["delta_rate_ab_ctrl"]),
            -x["abnormal_n"],
            x["gene"],
        )
    )
    return ranked


def gene_carrier_barh(ax, rows: list[dict], xlabel: str) -> None:
    import numpy as np

    if not rows:
        ax.text(0.5, 0.5, "No genes", ha="center", va="center", transform=ax.transAxes)
        return
    y = np.arange(len(rows))
    h = 0.24
    for j, g in enumerate(GROUPS):
        rates = [100 * fnum(r[f"{g}_rate"]) for r in rows]
        ax.barh(y + (1 - j) * h, rates, height=h, color=GROUP_COLOR[g], label=GROUP_LABEL[g], linewidth=0)
    ax.set_yticks(y)
    ax.set_yticklabels([r["gene"] for r in rows], fontstyle="italic")
    ax.set_xlabel(xlabel)
    ax.invert_yaxis()
    ax.legend(frameon=False, loc="lower right", fontsize=5)


def make_plots(out: Path) -> None:
    import matplotlib.pyplot as plt
    import numpy as np

    apply_nature_style()
    metrics = read_tsv(out / "sample_metrics.tsv")
    snv_gene = read_tsv(out / "snv_by_gene.tsv")
    sv_gene = read_tsv(out / "sv_by_gene.tsv")
    snv_long = read_tsv(out / "snv_variants.tsv")
    sv_long = read_tsv(out / "sv_variants.tsv")
    cards = read_tsv(out / "abnormal_sample_cards.tsv")

    by_g = {g: [r for r in metrics if r["Condition"] == g] for g in GROUPS}
    rare_counts = {g: [inum(r["n_snv_rare"]) for r in by_g[g]] for g in GROUPS}
    sv_counts = {g: [inum(r["n_sv_any"]) for r in by_g[g]] for g in GROUPS}
    exon_counts = {g: [inum(r["n_sv_exon"]) for r in by_g[g]] for g in GROUPS}

    n_groups = {g: len(by_g[g]) for g in GROUPS}
    sv_gene_rank = rank_genes_ab_vs_control(sv_gene, "any", n_groups)
    snv_gene_rank = rank_genes_ab_vs_control(snv_gene, "rare", n_groups)
    top_sv = sv_gene_rank[:12]
    top_snv = snv_gene_rank[:12]

    groups = {g: [r["Sample_ID"] for r in by_g[g]] for g in GROUPS}
    site_rows = sv_unique_site_rates(sv_long, groups)

    loc_order = ("exon_or_splice", "intron", "other", "unannotated")
    n_sv = {g: max(1, len([r for r in sv_long if r["Condition"] == g])) for g in GROUPS}
    src_loc = []
    for loc in loc_order:
        for g in GROUPS:
            k = sum(1 for r in sv_long if r["Condition"] == g and location_class(r.get("location", "")) == loc)
            src_loc.append({"group": g, "location_class": loc, "n": k, "pct": 100 * k / n_sv[g]})
    intron_pct = {
        r["group"]: r["pct"] for r in src_loc if r["location_class"] == "intron"
    }

    fig, axes = plt.subplots(2, 2, figsize=(180 * MM, 140 * MM))
    ax_a, ax_b, ax_sc, ax_d = axes.ravel()

    box_with_points(ax_a, rare_counts, "Rare nonsyn. SNV count\n(gnomAD AF < 1% or missing)", group_ns=n_groups)
    p_a = annotate_mw(ax_a, rare_counts)
    panel_label(ax_a, "a")
    box_with_points(ax_b, sv_counts, "Gene-overlapping SV count\n(any location, size-filtered)", group_ns=n_groups)
    p_b = annotate_mw(ax_b, sv_counts)
    panel_label(ax_b, "b")

    gene_carrier_barh(ax_sc, top_sv, "Carriers with panel SV (%)")
    panel_label(ax_sc, "c")

    gene_carrier_barh(ax_d, top_snv, "Carriers with rare SNV (%)")
    panel_label(ax_d, "d")

    fig.subplots_adjust(left=0.10, right=0.98, top=0.90, bottom=0.10, wspace=0.42, hspace=0.48)
    save_fig(fig, "fig1_epifactors_overview")
    plt.close(fig)

    write_source("fig1a_rare_snv_per_sample.tsv", metrics, list(metrics[0].keys()))
    write_source("fig3c_sv_gene_rank.tsv", sv_gene_rank, GENE_RANK_FIELDS)
    write_source("fig3d_snv_gene_rank.tsv", snv_gene_rank, GENE_RANK_FIELDS)
    write_source("fig1d_sv_location.tsv", src_loc, ["group", "location_class", "n", "pct"])
    write_source(
        "fig3c_sv_site_rates.tsv",
        site_rows,
        [
            "coord_key", "svtype", "acmg_class",
            "abnormal_n", "abnormal_rate", "normal_n", "normal_rate",
            "control_n", "control_rate", "p_abnormal_vs_control", "fdr_abnormal_vs_control",
        ],
    )
    write_source(
        "fig1ab_mannwhitney.tsv",
        [
            {"panel": "a_rare_snv", **{k: v for k, v in p_a.items()}},
            {"panel": "b_sv_any", **{k: v for k, v in p_b.items()}},
        ],
        ["panel", "abnormal_vs_normal", "abnormal_vs_control", "normal_vs_control"],
    )

    # Extended: consequence mix (abnormal) + exon SV counts + ACMG>=4 genes
    fig2, axes2 = plt.subplots(1, 3, figsize=(180 * MM, 62 * MM))
    ab_snv = [r for r in snv_long if r["Condition"] == "abnormal"]
    cons_order = ("missense", "lof", "splicing", "other_nonsyn")
    cons_n = [sum(1 for r in ab_snv if r.get("consequence") == c) for c in cons_order]
    axes2[0].bar(range(4), cons_n, color=["#0072B2", "#D55E00", "#E69F00", "#CC79A7"], width=0.7, linewidth=0)
    axes2[0].set_xticks(range(4))
    axes2[0].set_xticklabels(["Missense", "LoF", "Splice", "Other"], rotation=0)
    axes2[0].set_ylabel("Abnormal SNV events")
    panel_label(axes2[0], "a")

    box_with_points(axes2[1], exon_counts, "Exon/splice SV count")
    panel_label(axes2[1], "b")

    acmg_n = {g: sum(inum(r[f"acmg4_{g}_n"]) > 0 for r in sv_gene) for g in GROUPS}
    axes2[2].bar(
        range(3), [acmg_n[g] for g in GROUPS],
        color=[GROUP_COLOR[g] for g in GROUPS], width=0.6, linewidth=0,
    )
    axes2[2].set_xticks(range(3))
    axes2[2].set_xticklabels([GROUP_LABEL[g] for g in GROUPS])
    axes2[2].set_ylabel("Genes with ACMG ≥ 4 SV")
    panel_label(axes2[2], "c")
    fig2.subplots_adjust(left=0.08, right=0.98, top=0.86, bottom=0.18, wspace=0.45)
    save_fig(fig2, "fig2_consequence_exon_acmg")
    plt.close(fig2)
    write_source(
        "fig2a_consequence.tsv",
        [{"consequence": c, "n_abnormal": n} for c, n in zip(cons_order, cons_n)],
        ["consequence", "n_abnormal"],
    )

    n_ab = len(cards)
    n_rare_gene_ab = sum(1 for r in snv_gene if inum(r["rare_abnormal_n"]) > 0)
    n_higher = sum(
        1 for r in snv_gene
        if inum(r["rare_abnormal_n"]) > 0
        and fnum(r["rare_rate_ab_minus_normal"]) > 0
        and fnum(r["rare_rate_ab_minus_control"]) > 0
    )
    med = {g: float(np.median(rare_counts[g])) if rare_counts[g] else 0 for g in GROUPS}
    med_sv = {g: float(np.median(sv_counts[g])) if sv_counts[g] else 0 for g in GROUPS}
    acmg_ab = sum(inum(r["n_sv_acmg4"]) for r in cards)
    exon_ab = sum(inum(r["n_sv_exon"]) for r in cards)
    legend = PLOT_DIR / "figure_legends.txt"
    talk = PLOT_DIR / "talking_points.txt"
    legend.write_text(
        "Fig. 1 | EpiFactors-gene variants in methylation-defined groups (descriptive).\n"
        "a, Per-sample count of rare (gnomAD AF < 1% or missing) nonsynonymous or splice SNVs "
        f"in the EpiFactors library. n = {len(by_g['abnormal'])} abnormal, {len(by_g['normal'])} "
        f"normal, {len(by_g['control'])} control. Boxes show median and IQR; points are samples. "
        "b, Per-sample count of size-filtered SVs overlapping any EpiFactors gene (any location). "
        "c, Top EpiFactors genes by raw Fisher p (abnormal vs control SV carriers); bars are percent of samples. "
        "d, Top EpiFactors genes by raw Fisher p (abnormal vs control rare SNV carriers); same layout as c. "
        "Panel-overlapping SVs are mostly intronic "
        f"(abnormal {intron_pct.get('abnormal', 0):.0f}%, "
        f"normal {intron_pct.get('normal', 0):.0f}%, "
        f"control {intron_pct.get('control', 0):.0f}%); not shown as a panel. "
        f"a, two-sided Mann–Whitney: Ab–N p={fmt_p(p_a['abnormal_vs_normal'])}, "
        f"Ab–C p={fmt_p(p_a['abnormal_vs_control'])}, N–C p={fmt_p(p_a['normal_vs_control'])}. "
        f"b, Ab–N p={fmt_p(p_b['abnormal_vs_normal'])}, "
        f"Ab–C p={fmt_p(p_b['abnormal_vs_control'])}, N–C p={fmt_p(p_b['normal_vs_control'])}. "
        "d is descriptive (no test).\n\n"
        "Fig. 2 | SNV consequence mix and coding-relevant SVs.\n"
        "a, Consequence of nonsynonymous/splice SNVs in abnormal samples. "
        "b, Per-sample SVs whose AnnotSV Location includes exon or splice (including exonN-exonN). "
        "c, Number of genes carrying an overlapping SV with ACMG class ≥ 4.\n",
        encoding="utf-8",
    )
    talk.write_text(
        "汇报要点（描述性，不做组间显著性）\n"
        f"- 表型：abnormal {len(by_g['abnormal'])} / normal {len(by_g['normal'])} / "
        f"control {len(by_g['control'])}；基因库为完整 EpiFactors（EpiGenes_main.xlsx）。\n"
        f"- 47 例 abnormal 均有 panel 内非同义/剪接 SNV；每例 rare SNV 中位数约 {med['abnormal']:.0f} "
        f"（normal {med['normal']:.0f}，control {med['control']:.0f}）。\n"
        f"- Rare SNV 累及 {n_rare_gene_ab} 个基因；其中 {n_higher} 个基因 abnormal 比例同时高于 normal 与 control。\n"
        f"- Panel 重叠 SV 中位数约 {med_sv['abnormal']:.0f} / 人，多数为内含子；"
        f"abnormal 的 exon/splice SV 事件合计 {exon_ab}，ACMG≥4 事件合计 {acmg_ab}。\n"
        "- 高比例非同义 SNV（如接近 100%）多为共享常见错义，不宜解读为 abnormal 特异。\n"
        "- 图件：plots/epifactors/fig1_epifactors_overview.pdf（主图）与 fig2_*.pdf；"
        "源数据在 source_data/。\n"
        f"- 样本卡片：tables/epifactors/abnormal_sample_cards.tsv（n={n_ab}）。\n",
        encoding="utf-8",
    )
    print(f"  legends -> {legend}")
    print(f"  talking points -> {talk}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def tables_ready(out: Path) -> bool:
    return all((out / name).exists() for name in REQUIRED_LONG)


def maybe_adopt_split(out: Path) -> None:
    old = MODULE / "tables" / "epifactors_split"
    if tables_ready(out) or not old.exists():
        return
    out.mkdir(parents=True, exist_ok=True)
    for p in old.glob("*.tsv"):
        shutil.copy2(p, out / p.name)
    print(f"copied previous split tables -> {out}")


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--catalog", action="store_true", help="rescan ANNOVAR, VCF, AnnotSV")
    p.add_argument("--plot-only", action="store_true", help="tables + figures from existing long TSVs")
    p.add_argument(
        "--gw8-10",
        action="store_true",
        help="restrict derived tables/plots to gestational weeks 8–10; do not overwrite 648-wide longs",
    )
    p.add_argument("--from-longs", type=Path, default=None, help="directory with snv_variants.tsv / sv_variants.tsv")
    p.add_argument("--derived-out", type=Path, default=None, help="write remapped derived tables here")
    p.add_argument("--pheno", type=Path, default=None, help="phenotype TSV (VCF_Sample_ID, Condition)")
    args = p.parse_args()

    src = Path(args.from_longs) if args.from_longs else OUT_DIR
    src.mkdir(parents=True, exist_ok=True)
    if args.from_longs is None:
        maybe_adopt_split(src)

    keep_ids: set[str] | None = None
    out = Path(args.derived_out) if args.derived_out else src
    pheno_path = Path(args.pheno) if args.pheno else PHENO_PATH
    if args.gw8_10:
        keep_ids = gw8_10_sample_ids(load_phenotype(pheno_path))
        out = MODULE / "tables" / "gw8_10" / "epifactors"
        out.mkdir(parents=True, exist_ok=True)
        print(f"subset gestational weeks 8–10: {len(keep_ids)} samples -> {out}")
    elif args.derived_out:
        out.mkdir(parents=True, exist_ok=True)

    do_catalog = args.catalog or (not args.plot_only and not tables_ready(src))
    if do_catalog:
        if not EPIGENES_XLSX.exists():
            raise SystemExit(f"missing {EPIGENES_XLSX}")
        panel_rows = load_panel_from_xlsx(EPIGENES_XLSX)
        snv_long, sv_long = run_catalog(src, panel_rows)
    else:
        print(f"using existing tables in {src}")
        panel_rows = read_tsv(src / "gene_panel.tsv")
        set_panel(panel_rows)
        snv_long = read_tsv(src / "snv_variants.tsv")
        sv_long = read_tsv(src / "sv_variants.tsv")
        for r in sv_long:
            r["exon_or_splice"] = sv_exon_or_splice(r.get("location", ""))
        if not args.gw8_10 and args.derived_out is None and args.pheno is None:
            write_tsv(
                src / "sv_variants.tsv",
                sv_long,
                [
                    "Sample_ID", "Condition", "gene", "gene_class", "chrom", "start", "end",
                    "variant_pos", "svtype", "svlen", "sv_id", "gt", "location",
                    "exon_or_splice", "acmg_class", "acmg_ge4",
                ],
            )

    pheno_map = {r["VCF_Sample_ID"]: r["Condition"] for r in load_phenotype(pheno_path)}
    if args.pheno is not None:
        for r in snv_long:
            if r.get("Sample_ID") in pheno_map:
                r["Condition"] = pheno_map[r["Sample_ID"]]
        for r in sv_long:
            if r.get("Sample_ID") in pheno_map:
                r["Condition"] = pheno_map[r["Sample_ID"]]

    if keep_ids is not None:
        snv_long = [r for r in snv_long if r["Sample_ID"] in keep_ids]
        sv_long = [r for r in sv_long if r["Sample_ID"] in keep_ids]
        shutil.copy2(src / "gene_panel.tsv", out / "gene_panel.tsv")
        if snv_long:
            write_tsv(out / "snv_variants.tsv", snv_long, list(snv_long[0].keys()))
        else:
            write_tsv(out / "snv_variants.tsv", [], ["Sample_ID"])
        sv_fields = [
            "Sample_ID", "Condition", "gene", "gene_class", "chrom", "start", "end",
            "variant_pos", "svtype", "svlen", "sv_id", "gt", "location",
            "exon_or_splice", "acmg_class", "acmg_ge4",
        ]
        write_tsv(out / "sv_variants.tsv", sv_long, sv_fields)
    elif args.derived_out is not None:
        shutil.copy2(src / "gene_panel.tsv", out / "gene_panel.tsv")
        if snv_long:
            write_tsv(out / "snv_variants.tsv", snv_long, list(snv_long[0].keys()))
        sv_fields = [
            "Sample_ID", "Condition", "gene", "gene_class", "chrom", "start", "end",
            "variant_pos", "svtype", "svlen", "sv_id", "gt", "location",
            "exon_or_splice", "acmg_class", "acmg_ge4",
        ]
        write_tsv(out / "sv_variants.tsv", sv_long, sv_fields)

    refresh_derived(out, snv_long, sv_long, keep_ids=keep_ids, pheno_path=pheno_path)
    print("[plot]")
    make_plots(out)
    print(f"tables: {out}")
    print(f"plots:  {PLOT_DIR}")


if __name__ == "__main__":
    main()
